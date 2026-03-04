"""Excel output generation.

Writes transcription results into the provided template Excel file,
populating Column C (Transcription) for each participant sheet.
"""

from __future__ import annotations

import copy
import logging
from pathlib import Path
from typing import Any, Dict, List, Optional

import openpyxl

from .config import PipelineConfig, AudioFileConfig
from .experiment_logger import ParticipantResult

logger = logging.getLogger(__name__)


# Map sheet_name in config → sheet name in Excel
SHEET_MAP = {
    "38010-2A": "38010-2A",
    "38011-1A": "38011-1A",
    "38012-2A": "38012-2A",
    "38015-1A": "38015-1A",
}


def write_results_to_excel(
    results: List[ParticipantResult],
    template_path: str,
    output_path: str,
) -> str:
    """Write transcription results to the Excel template.

    Parameters
    ----------
    results : List[ParticipantResult]
        Transcription results for each participant.
    template_path : str
        Path to the template Excel file.
    output_path : str
        Path where the output Excel file will be saved.

    Returns
    -------
    str
        Path to the saved Excel file.
    """
    wb = openpyxl.load_workbook(template_path)

    for participant in results:
        sheet_name = participant.sheet_name

        if sheet_name not in wb.sheetnames:
            logger.warning(
                "Sheet '%s' not found in template. Available: %s",
                sheet_name, wb.sheetnames,
            )
            continue

        ws = wb[sheet_name]
        logger.info(
            "Writing %d transcriptions to sheet '%s'",
            len(participant.transcriptions), sheet_name,
        )

        for trans in participant.transcriptions:
            # Sentence numbers are 1-based, data starts at row 2
            row = trans.sentence_number + 1  # row 1 is header

            # Column C = Transcription
            cell = ws.cell(row=row, column=3)
            cell.value = trans.transcription

            # If flagged, add a comment or highlight
            if trans.flagged:
                # Use column E for flags (keeping D for scoring)
                flag_cell = ws.cell(row=row, column=5)
                flag_cell.value = f"[FLAGGED: {trans.flag_reason}]"

                # Light yellow background for flagged cells
                from openpyxl.styles import PatternFill
                cell.fill = PatternFill(
                    start_color="FFFF99", end_color="FFFF99", fill_type="solid"
                )

    # Save
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    logger.info("Saved output Excel: %s", output_path)

    return output_path


def write_detailed_results(
    results: List[ParticipantResult],
    output_path: str,
) -> str:
    """Write a detailed results Excel with confidence scores and metadata.

    Creates a new workbook with:
    - Summary sheet
    - Per-participant sheets with all metadata
    """
    wb = openpyxl.Workbook()

    # Summary sheet
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary.append(["Participant", "Version", "Segments", "Avg Confidence",
                        "Flagged", "No Response", "Processing Time (s)"])

    for p in results:
        ws_summary.append([
            p.participant_id,
            p.eit_version,
            p.segments_found,
            round(p.avg_confidence, 4),
            p.low_confidence_count,
            p.no_response_count,
            round(p.total_processing_time_s, 1),
        ])

    # Per-participant detailed sheets
    for p in results:
        ws = wb.create_sheet(title=f"{p.participant_id}-{p.eit_version}")
        ws.append([
            "Sentence #", "Stimulus", "Transcription", "Raw Transcription",
            "Avg Log Prob", "No Speech Prob", "Flagged", "Flag Reason",
            "Segment Start (s)", "Segment End (s)", "Pass #",
        ])

        for t in p.transcriptions:
            ws.append([
                t.sentence_number,
                t.stimulus,
                t.transcription,
                t.raw_transcription,
                round(t.avg_log_prob, 4),
                round(t.no_speech_prob, 4),
                t.flagged,
                t.flag_reason,
                round(t.segment_start_s, 2),
                round(t.segment_end_s, 2),
                t.pass_number,
            ])

    # Auto-adjust column widths
    for ws in wb.worksheets:
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max_length + 2, 60)

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    logger.info("Saved detailed results: %s", output_path)
    return output_path
