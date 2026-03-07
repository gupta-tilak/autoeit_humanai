"""Stage 9 — Final Output.

Assembles structured dataset from all pipeline stages and exports
to JSON, CSV, and Excel formats.
"""

from __future__ import annotations

import csv
import json
import logging
from dataclasses import asdict, dataclass, field
from pathlib import Path
from typing import Any, Dict, List, Optional

from .asr_transcription import TranscriptionResult
from .config import OutputConfig, TARGET_SENTENCES
from .leakage_check import LeakageFlag
from .response_window import ResponseSegment

logger = logging.getLogger("eit_pipeline.output")


@dataclass
class SegmentResult:
    """Complete result for a single EIT segment."""
    sentence_id: int
    stimulus_text: str
    stimulus_start_s: float
    stimulus_end_s: float
    response_start_s: float
    response_end_s: float
    transcript: str
    confidence: float
    avg_log_prob: float = 0.0
    no_speech_prob: float = 0.0
    flagged: bool = False
    flag_reason: str = ""
    processing_time_s: float = 0.0


@dataclass
class ParticipantOutput:
    """Complete output for one participant."""
    participant_id: str
    eit_version: str
    segments: List[SegmentResult]
    metadata: Dict[str, Any] = field(default_factory=dict)


def assemble_results(
    participant_id: str,
    eit_version: str,
    responses: List[ResponseSegment],
    transcriptions: List[TranscriptionResult],
    leakage_flags: List[LeakageFlag],
    stimulus_sentences: Optional[List[str]] = None,
) -> ParticipantOutput:
    """Assemble results from all pipeline stages into a unified output.

    Parameters
    ----------
    participant_id : str
        Participant identifier.
    eit_version : str
        EIT version (1A or 2A).
    responses : list of ResponseSegment
        Response windows.
    transcriptions : list of TranscriptionResult
        ASR transcription results.
    leakage_flags : list of LeakageFlag
        Leakage check results.
    stimulus_sentences : list of str, optional
        Reference sentences.

    Returns
    -------
    ParticipantOutput
    """
    if stimulus_sentences is None:
        stimulus_sentences = TARGET_SENTENCES

    # Index transcriptions and flags by sentence_id
    trans_by_id = {t.sentence_id: t for t in transcriptions}
    flags_by_id = {f.sentence_id: f for f in leakage_flags}

    segments = []
    for resp in responses:
        sid = resp.sentence_id
        stim_text = stimulus_sentences[sid - 1] if 1 <= sid <= len(stimulus_sentences) else ""

        trans = trans_by_id.get(sid)
        flag = flags_by_id.get(sid)

        transcript = trans.transcript if trans else "[no response]"
        confidence = resp.confidence
        avg_log_prob = trans.avg_log_prob if trans else 0.0
        no_speech_prob = trans.no_speech_prob if trans else 0.0
        processing_time = trans.processing_time_s if trans else 0.0

        flagged = flag.flagged if flag else False
        flag_reason = flag.reason if flag else ""

        segments.append(SegmentResult(
            sentence_id=sid,
            stimulus_text=stim_text,
            stimulus_start_s=resp.stimulus_start_s,
            stimulus_end_s=resp.stimulus_end_s,
            response_start_s=resp.response_start_s,
            response_end_s=resp.response_end_s,
            transcript=transcript,
            confidence=confidence,
            avg_log_prob=avg_log_prob,
            no_speech_prob=no_speech_prob,
            flagged=flagged,
            flag_reason=flag_reason,
            processing_time_s=processing_time,
        ))

    return ParticipantOutput(
        participant_id=participant_id,
        eit_version=eit_version,
        segments=segments,
    )


def export_json(
    output: ParticipantOutput,
    path: str,
) -> None:
    """Export results to JSON."""
    path_obj = Path(path)
    path_obj.parent.mkdir(parents=True, exist_ok=True)

    data = {
        "participant_id": output.participant_id,
        "eit_version": output.eit_version,
        "total_segments": len(output.segments),
        "segments_with_response": sum(
            1 for s in output.segments if s.response_end_s > 0
        ),
        "flagged_count": sum(1 for s in output.segments if s.flagged),
        "metadata": output.metadata,
        "segments": [asdict(s) for s in output.segments],
    }

    path_obj.write_text(
        json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8",
    )
    logger.info("Exported JSON: %s", path)


def export_csv(
    output: ParticipantOutput,
    path: str,
) -> None:
    """Export results to CSV."""
    path_obj = Path(path)
    path_obj.parent.mkdir(parents=True, exist_ok=True)

    fieldnames = [
        "sentence_id", "stimulus_text",
        "stimulus_start_s", "stimulus_end_s",
        "response_start_s", "response_end_s",
        "transcript", "confidence",
        "avg_log_prob", "no_speech_prob",
        "flagged", "flag_reason",
    ]

    with open(path_obj, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for seg in output.segments:
            row = asdict(seg)
            # Keep only the fields we want
            writer.writerow({k: row[k] for k in fieldnames})

    logger.info("Exported CSV: %s", path)


def export_excel(
    output: ParticipantOutput,
    path: str,
) -> None:
    """Export results to Excel."""
    try:
        import openpyxl
    except ImportError:
        logger.warning("openpyxl not installed — skipping Excel export")
        return

    path_obj = Path(path)
    path_obj.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{output.participant_id}-{output.eit_version}"

    # Header
    headers = [
        "Sentence #", "Stimulus", "Transcript",
        "Response Start (s)", "Response End (s)",
        "Confidence", "Avg Log Prob",
        "Flagged", "Flag Reason",
    ]
    ws.append(headers)

    # Bold header
    from openpyxl.styles import Font
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # Data rows
    for seg in output.segments:
        row = [
            seg.sentence_id,
            seg.stimulus_text,
            seg.transcript,
            round(seg.response_start_s, 2),
            round(seg.response_end_s, 2),
            round(seg.confidence, 3),
            round(seg.avg_log_prob, 3),
            seg.flagged,
            seg.flag_reason,
        ]
        ws.append(row)

        # Highlight flagged rows
        if seg.flagged:
            from openpyxl.styles import PatternFill
            yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            for cell in ws[ws.max_row]:
                cell.fill = yellow

    # Auto-width columns
    for col in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)

    wb.save(str(path_obj))
    logger.info("Exported Excel: %s", path)


def export_all(
    output: ParticipantOutput,
    output_dir: str,
    config: Optional[OutputConfig] = None,
) -> Dict[str, str]:
    """Export results in all configured formats.

    Returns dict of format -> output path.
    """
    if config is None:
        config = OutputConfig()

    out_dir = Path(output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    prefix = f"{output.participant_id}_{output.eit_version}"
    paths = {}

    if config.export_json:
        p = str(out_dir / f"{prefix}_results.json")
        export_json(output, p)
        paths["json"] = p

    if config.export_csv:
        p = str(out_dir / f"{prefix}_results.csv")
        export_csv(output, p)
        paths["csv"] = p

    if config.export_excel:
        p = str(out_dir / f"{prefix}_results.xlsx")
        export_excel(output, p)
        paths["excel"] = p

    return paths
